from dotenv import load_dotenv

from flask import Flask, jsonify, request, send_from_directory, Response, stream_with_context
from flask_cors import CORS
from flasgger import Swagger

import mysql.connector
import pandas as pd
import csv, decimal, io, os, json
from datetime import datetime, date

from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill

from insertarMedicionV2 import insertar_medicion_bp as insertar_medicion_v2_bp
from alertas import alertas_bp
from files import files_bp
# from flask_socketio import SocketIO, emit




load_dotenv()
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'  # Reemplaza con una clave secreta segura
app.config['SWAGGER'] = {
    'title': 'API Docs - Sensores',
    'description': 'Esta es la documentación interactiva para la API. Incluye detalles sobre los endpoints disponibles, sus parámetros, y ejemplos de uso.',
    'uiversion': 3,
    'favicon': 'https://example.com/favicon.ico',  # URL de tu favicon personalizado
    'specs': [
        {
            'endpoint': 'apispec_1',
            'route': '/apispec_1.json',
            'rule_filter': lambda rule: True,  # Todos los endpoints están documentados
            'model_filter': lambda tag: True,  # Todos los modelos están incluidos
        }
    ],
    'static_url_path': '/flasgger_static',
    'swagger_ui': True,
    'specs_route': '/apidocs/',  # URL de acceso a la documentación
    'contact': {
        'name': 'Soporte API',
        'url': 'https://example.com/soporte',
        'email': 'soporte@example.com'
    },
    'license': {
        'name': 'MIT License',
        'url': 'https://opensource.org/licenses/MIT'
    },
    'servers': [
        {
            'url': 'http://localhost:8084',
            'description': 'Servidor local de desarrollo'
        },
        {
            'url': 'https://api-sensores.cmasccp.cl',
            'description': 'Servidor de producción'
        }
    ],
    'tags': [
        {'name': 'Datos', 'description': 'Endpoints relacionados con la manipulación de datos.'},
        {'name': 'Sensores', 'description': 'Endpoints relacionados con la gestión de sensores.'},
        {'name': 'Tablas', 'description': 'Endpoints para operaciones sobre tablas.'},
        {'name': 'Esquemas', 'description': 'Endpoints para obtener esquemas de tablas.'}
    ]
}

CORS(app,
     resources={r"/*": {"origins": ["https://sensores.cmasccp.cl", "https://api-sensores.cmasccp.cl", "http://localhost:5173"]}},
     supports_credentials=True,
     allow_headers=["Content-Type", "Authorization"],
     expose_headers=["Content-Type"],
     methods=["GET","POST","PUT","DELETE","OPTIONS"])
swagger = Swagger(app)

@app.after_request
def add_cors_headers(response):
    """Asegura cabeceras CORS para orígenes permitidos."""
    origin = request.headers.get('Origin')
    allowed = ["https://sensores.cmasccp.cl", "https://api-sensores.cmasccp.cl", "http://localhost:5173"]
    if origin in allowed:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
    # Cabeceras generales que ayudan en preflight
    response.headers.setdefault('Access-Control-Allow-Headers', 'Content-Type, Authorization')
    response.headers.setdefault('Access-Control-Allow-Methods', 'GET,POST,PUT,DELETE,OPTIONS')
    return response


ALLOWED_TABLES_PROP = [
{'displayName':'Datos','dataName':'datos'},
{'displayName':'Dispositivos','dataName':'dispositivos'},
{'displayName':'Estados','dataName':'estados'},
{'displayName':'Grupos','dataName':'grupos'},
{'displayName':'Personas','dataName':'personas'},
{'displayName':'Proyectos','dataName':'proyectos'},
{'displayName':'Roles','dataName':'roles'},
{'displayName':'Roles en grupos','dataName':'roles_en_grupos'},
{'displayName':'Roles en proyectos','dataName':'roles_en_proyectos'},
{'displayName':'Sensores','dataName':'sensores'},
{'displayName':'Sensores en dispositivo','dataName':'sensores_en_dispositivo'},
{'displayName':'Sensores tipo','dataName':'sensores_tipo'},
{'displayName':'Sesiones','dataName':'sesiones'},
{'displayName':'Variables','dataName':'variables'},
{'displayName':'Variables en sensores','dataName':'variables_en_sensores'},
{'displayName':'Imágenes','dataName':'imagenes'},
]


FOREIGN_KEYS_PROP = {
    "id_sesion": {"table": "sesiones","columns": ["id_sesion", "descripcion"]}, 
    "id_variable": {"table": "variables", "columns":["id_variable", "descripcion"]}, 
    "idVariable": {"table": "variables", "columns":["id_variable", "descripcion"]}, 
    "id_grupo": {"table": "grupos", "columns":["id_grupo", "nombre"]}, 
    "id_estado": {"table": "estados", "columns":["id_estado", "nombre"]}, 
    "id_proyecto": {"table": "proyectos", "columns":["id_proyecto", "nombre"]}, 
    "id_persona": {"table": "personas", "columns":["id_persona", "nombre", "apellido"]}, 
    "id_persona_responsable": {"table": "personas", "columns":["id_persona", "nombre", "apellido"]},
    "id_persona_responsable_ingreso": {"table": "personas", "columns":["id_persona", "nombre", "apellido"]}, 
    "id_persona_responsable_salida": {"table": "personas", "columns":["id_persona", "nombre", "apellido"]}, 
    "id_sensor": {"table": "sensores", "columns":["id_sensor", "numero_serial"]}, 
    "id_sensor_tipo": {"table": "sensores_tipo", "columns":["id_sensor_tipo", "marca", "modelo"]},
    "idSensorTipo": {"table": "sensores_tipo", "columns":["id_sensor_tipo", "marca", "modelo"]},
    "id_rol": {"table": "roles", "columns":["id_rol", "nombre"]},
    "id_dispositivo": {"table": "dispositivos", "columns":["id_dispositivo", "codigo_interno"]},
    }

ALLOWED_TABLES = [table['dataName'] for table in ALLOWED_TABLES_PROP]
config = {
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
    "host": os.getenv("DB_HOST"),
    "database": os.getenv("DB_NAME"),
    "port": int(os.getenv("DB_PORT", 3306)),  # Valor por defecto: 3306
}
print(config)

# socketio = SocketIO(app, cors_allowed_origins="*", logger=True, engineio_logger=True)

# @socketio.on('connect')
# def handle_connect():
#     print('Cliente conectado')
#     emit('message', {'data': 'Conexión exitosa'})

# @socketio.on('disconnect')
# def handle_disconnect():
#     print('Cliente desconectado')

# @socketio.on_error_default
# def default_error_handler(e):
#     print(f'SocketIO Error: {e}')

# @app.route('/test-websocket', methods=['GET'])
# def test_websocket():
#     """
#     Endpoint de prueba para verificar que WebSocket funciona correctamente.
#     ---
#     tags:
#       - Testing
#     responses:
#       200:
#         description: Mensaje de prueba enviado por WebSocket
#     """
#     try:
#         test_data = {
#             'message': 'Test message from server',
#             'timestamp': datetime.now().isoformat()
#         }
#         socketio.emit('test_message', test_data)
#         return jsonify({
#             'status': 'success', 
#             'message': 'Test WebSocket message sent',
#             'data': test_data
#         }), 200
#     except Exception as e:
#         return jsonify({
#             'status': 'error',
#             'message': f'WebSocket test failed: {str(e)}'
#         }), 500


app.register_blueprint(insertar_medicion_v2_bp)
app.register_blueprint(alertas_bp)
app.register_blueprint(files_bp)


@app.route('/endovenosaDummy', methods=['GET'])
def endovenosa_dummy():
    """
    Devuelve datos simulados de un dispositivo.
    ---
    tags:
      - Simulaciones
    responses:
      200:
        description: Datos del dispositivo
        examples:
          application/json: {
            "name": "Dispositivo 1",
            "license": "JLZJ41",
            "password": "3508239",
            "firmwareVersion": "v10.3",
            "status": "Transmitting",
            "lastConnection": "07/10/2024",
            "alertMsg": "Burbuja de aire detectada",
            "alertType": "Danger"
          }
    """
    return jsonify({
        'name': "Dispositivo 1",
        'license': "JLZJ41",
        'password': "3508239",
        'firmwareVersion': "v10.3",
        'status': "Transmitting",
        'lastConnection': "07/10/2024",
        'alertMsg': "Burbuja de aire detectada",
        'alertType': "Danger"}), 200

@app.route('/ultimoValor', methods=['GET'])
def ultimo_valor():
    """
    Obtiene el último valor de una columna específica de una tabla.
    ---
    tags:
      - Consultas
    parameters:
      - name: tabla
        in: query
        type: string
        required: true
        description: Nombre de la tabla desde la que se desea obtener el valor.
      - name: columna
        in: query
        type: string
        required: true
        description: Nombre de la columna de la que se desea obtener el último valor.
    responses:
      200:
        description: Último valor obtenido con éxito.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            data:
              type: string
              example: "123.45"
      400:
        description: Faltan parámetros requeridos (tabla o columna).
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "Debe proporcionar el nombre de la tabla y la columna"
      404:
        description: No se encontraron resultados en la tabla.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "No se encontraron resultados"
      500:
        description: Error interno en la base de datos o inesperado.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "Error en la base de datos: Error específico"
    """
    tabla = request.args.get('tabla')
    columna = request.args.get('columna')

    if not tabla or not columna:
        return jsonify({'status': 'fail', 'error': 'Debe proporcionar el nombre de la tabla y la columna'}), 400

    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        query = f"SELECT {columna} FROM {tabla} ORDER BY {columna} DESC LIMIT 1"
        cursor.execute(query)
        resultado = cursor.fetchone()

        if resultado:
            ultimo_valor = resultado[0]
            return jsonify({'status': 'success', 'data': ultimo_valor}), 200
        else:
            return jsonify({'status': 'fail', 'error': 'No se encontraron resultados'}), 404

    except mysql.connector.Error as e:
        return jsonify({'status': 'fail', 'error': f'Error en la base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'status': 'fail', 'error': f'Error inesperado: {str(e)}'}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/generarSesion', methods=['GET'])
def generar_sesion():
    """
    Crea una nueva sesión en la base de datos.
    ---
    tags:
      - Sesiones
    parameters:
      - name: id_proyecto
        in: query
        type: integer
        required: true
        description: ID del proyecto al que pertenece la sesión.
      - name: id_persona_responsable
        in: query
        type: integer
        required: false
        description: ID de la persona responsable de la sesión. Opcional.
      - name: descripcion
        in: query
        type: string
        required: false
        description: Descripción de la sesión. Predeterminado a una cadena vacía.
      - name: fecha_inicio
        in: query
        type: string
        format: datetime
        required: false
        description: Fecha de inicio de la sesión en formato "YYYY-MM-DD HH:MM:SS". Predeterminado a la fecha actual.
      - name: version
        in: query
        type: string
        required: false
        description: Versión de la sesión. Predeterminado a "1.0".
      - name: ubicacion
        in: query
        type: string
        required: false
        description: Ubicación de la sesión. Predeterminado a una cadena vacía.
    responses:
      201:
        description: Sesión creada correctamente.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            message:
              type: string
              example: Sesión creada correctamente
            generated_id:
              type: integer
              example: 123
      400:
        description: Parámetro obligatorio "id_proyecto" no proporcionado.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: El parámetro id_proyecto es obligatorio
      500:
        description: Error interno en el servidor o base de datos.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Error interno
    """
    id_proyecto = request.args.get('id_proyecto')  # Obligatorio
    id_persona_responsable = request.args.get('id_persona_responsable')  # Opcional
    descripcion = request.args.get('descripcion', '')  # Predeterminado a cadena vacía
    fecha_inicio = request.args.get('fecha_inicio', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))  # Predeterminado a la fecha actual
    version = request.args.get('version', '1.0')  # Predeterminado a "1.0"
    ubicacion = request.args.get('ubicacion', '')  # Predeterminado a cadena vacía

    # Validar parámetros obligatorios
    if not id_proyecto:
        return jsonify({'status': 'fail', 'error': 'El parámetro id_proyecto es obligatorio'}), 400

    valores = [
        None,
        id_proyecto,
        id_persona_responsable if id_persona_responsable else None,  # Si no se proporciona, usar NULL
        descripcion,
        fecha_inicio,
        version,
        ubicacion
    ]

    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        sql_query = """
            INSERT INTO sesiones (id_sesion, id_proyecto, id_persona_responsable, descripcion, fecha_inicio, version, ubicacion)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """

        cursor.execute(sql_query, valores)
        conn.commit()
        generated_id = cursor.lastrowid

        return jsonify({
            'status': 'success',
            'message': 'Sesión creada correctamente',
            'generated_id': generated_id
        }), 201
    except Exception as e:
        return jsonify({'status': 'fail', 'error': str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/columnaForanea', methods=['GET'])
def columna_foranea():
    """
    Obtiene los valores relacionados con una columna foránea específica.
    ---
    tags:
      - Relaciones
    parameters:
      - name: columna
        in: query
        type: string
        required: true
        description: Nombre de la columna foránea para obtener los datos relacionados.
    responses:
      200:
        description: Datos obtenidos con éxito.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            data:
              type: array
              items:
                type: object
                properties:
                  value:
                    type: string
                    example: "1"
                  label:
                    type: string
                    example: "Nombre - Apellido"
      400:
        description: No se encontraron datos relacionados con la columna proporcionada.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: No se han obtenido los datos
      500:
        description: Error en la base de datos o error interno del servidor.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Error al conectarse a la base de datos <detalle del error>
    """
    args = request.args
    column = args.get('columna')
    try:
        print("Antes")
        conn = mysql.connector.connect(**config)
        print("Despues")
        cursor = conn.cursor(dictionary=True)

        if column in FOREIGN_KEYS_PROP.keys():
            table_name = FOREIGN_KEYS_PROP[column]["table"]
            columns = FOREIGN_KEYS_PROP[column]["columns"]

            columnas_str = ", ".join(columns)
            query = f"SELECT {columnas_str} FROM {table_name}"

            cursor.execute(query)
            filas = cursor.fetchall()

            transformed_data = [
                {
                    "value": fila[columns[0]],  # El valor de la columna principal
                    "label": " - ".join(str(fila[col]) for col in columns if col != column)  # Concatenar otras columnas
                }
                for fila in filas
            ]

            return jsonify({"status": "success", "data": transformed_data}), 200

        return jsonify({"status": "fail", "error": "No se han obtenido los datos"}), 400
    
    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"
        print(mensaje_error)
        return jsonify({"status": "fail", "error": mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({"status": "fail", "error": mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()         


@app.route('/insertarMedicion', methods=['GET'])
def insertar_medicion():
    """
    Inserta mediciones asociadas a sensores, variables y sesiones.
    ---
    tags:
      - Datos
    parameters:
      - name: times
        in: query
        type: string
        required: false
        description: Tiempos de las mediciones en formato de timestamp, separados por comas. Si no se especifica, se utilizará el tiempo actual.
      - name: idsSesiones
        in: query
        type: string
        required: false
        description: IDs de las sesiones asociadas a las mediciones, separados por comas. Si no se especifica, se utilizará NULL.
      - name: idsSensores
        in: query
        type: string
        required: true
        description: IDs de los sensores, separados por comas.
      - name: idsVariables
        in: query
        type: string
        required: true
        description: IDs de las variables, separados por comas.
      - name: valores
        in: query
        type: string
        required: true
        description: Valores de las mediciones, separados por comas.
    responses:
      201:
        description: Mediciones insertadas correctamente.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            message:
              type: string
              example: Registro insertado correctamente
      400:
        description: Las longitudes de los parámetros no coinciden.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Las longitudes de los parámetros no coinciden
      500:
        description: Error en la base de datos o error interno.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Error al conectarse a la base de datos <detalle del error>
    """

    timestamps = request.args.get('times', '').split(',')
    sesiones_ids = request.args.get('idsSesiones', '').split(',')
    sensor_ids = request.args.get('idsSensores', '').split(',')
    variable_ids = request.args.get('idsVariables', '').split(',')
    values = request.args.get('valores', '').split(',')

    # Si timestamps tiene un solo valor
    # se entrega un UNIXTIME  
    if len(timestamps) == 1 and timestamps[0]:  # Un solo valor
        timestamps = [timestamps[0]] * len(sensor_ids)

    # Si timestamps NO tiene valor
    elif not timestamps[0]:  
        current_timestamp = datetime.now().timestamp()
        timestamps = [str(current_timestamp)] * len(sensor_ids)

    # Si sesiones_ids tiene un solo valor
    if len(sesiones_ids) == 1 and sesiones_ids[0]:  # Un solo valor
        sesiones_ids = [sesiones_ids[0]] * len(sensor_ids)

    # Si sesiones_ids NO tiene valor
    elif not sesiones_ids[0]:  
        sesiones_ids = [None] * len(sensor_ids)


    if not (len(timestamps) == len(sensor_ids) == len(variable_ids) == len(values) == len(sesiones_ids)):
        return jsonify({'status': 'fail', 'error': 'Las longitudes de los parametros no coinciden'}), 400
    
    measurements = []

    # Fecha de inserción (servidor)  
    server_timestamp = datetime.now().timestamp()
    insertion_timestamps = [str(server_timestamp)] * len(sensor_ids)


    for i in range(len(sensor_ids)):
        timestamp_float = float(timestamps[i])
        datetime_obj = datetime.fromtimestamp(timestamp_float)
        formatted_datetime = datetime_obj.strftime('%Y-%m-%d %H:%M:%S')
        # fechas insercion
        insertion_timestamp_float = float(insertion_timestamps[i])
        insertion_datetime_obj = datetime.fromtimestamp(insertion_timestamp_float)
        formatted_insertion_datetime = insertion_datetime_obj.strftime('%Y-%m-%d %H:%M:%S')

        # print("Formatted datetime:", formatted_datetime)
        print("Formatted insertion datetime:", formatted_insertion_datetime)
        # Convertir valores vacíos a None para insertarlos como NULL en la base de datos
        valor = values[i] if values[i] and values[i].strip() else None

        measurements.append({
            "timestamp":formatted_datetime, #timestamps[i],
            "sesionId": sesiones_ids[i],
            "sensorId": sensor_ids[i],
            "variableId": variable_ids[i],
            "value": valor,
            "insertionTimestamp": formatted_insertion_datetime
        })

    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        for measurement in measurements:
            valores = [measurement['sensorId'], measurement['value'], measurement['timestamp'], measurement['variableId'], measurement['sesionId'], measurement['insertionTimestamp']]
            sql_query = f"INSERT INTO datos (id_sensor, valor, fecha, id_variable, id_sesion, fecha_insercion) VALUES (%s, %s, %s, %s, %s, %s)"
            log_query = sql_query % tuple(valores)  # Para fines de depuración
            print("Consulta SQL para depuración:", log_query)
            cursor.execute(sql_query, valores)

        conn.commit()
        # data_websocket = {"dispositivoId": "Desconocido", "fecha": formatted_datetime, "sesionesIds": sesiones_ids, "sensorIds": sensor_ids, "variableIds": variable_ids, "valores": values}
        # data_websocket = []
        # for i, measurement in enumerate(measurements):
        #     m = measurement.copy()
        #     m["dispositivoId"] = dispositivo_id[i] if i < len(dispositivo_id) else dispositivo_id[0]
        #     data_websocket.append(m)
        
        
        # Emitir mensaje por SocketIO después del commit exitoso
        # try:
        #     socketio.emit('medicion_insertada', data_websocket)
        #     print("WebSocket message emitted successfully")
        # except Exception as ws_error:
        #     print(f"Error emitting WebSocket message: {ws_error}")
        #     # No fallar la operación si WebSocket falla

        return jsonify({'status': 'success', 'message': 'Registro insertado correctamente'}), 201, {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'}

    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/listarTablas', methods=['GET'])
def listar_tablas():
    """
    Lista las tablas permitidas en la base de datos.
    ---
    tags:
      - Tablas
    responses:
      200:
        description: Lista de tablas obtenida con éxito.
        schema:
          type: array
          items:
            type: object
            properties:
              displayName:
                type: string
                example: "Datos"
              dataName:
                type: string
                example: "datos"
      500:
        description: Error interno en el servidor.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Error interno del servidor
    """
    return jsonify(ALLOWED_TABLES_PROP)

@app.route('/listarDatos', methods=['GET'])
def listar_datos():
    """
    Lista los datos de una tabla permitida con filtros opcionales.
    ---
    tags:
      - Tablas
    parameters:
      - name: tabla
        in: query
        type: string
        required: true
        description: Nombre de la tabla desde donde se obtendrán los datos.
      - name: limite
        in: query
        type: integer
        required: false
        description: Número máximo de registros a retornar. Sin límite si no se especifica.
      - name: offset
        in: query
        type: integer
        required: false
        description: Desplazamiento inicial para la consulta. Predeterminado a 0.
      - name: formato
        in: query
        type: string
        required: false
        description: Formato de salida 'json' o 'csv'. Predeterminado a 'json'.
      - name: filtros
        in: query
        type: string
        required: false
        description: Filtros opcionales para columnas específicas en la forma 'columna=valor1,valor2'.
    responses:
      200:
        description: Datos obtenidos con éxito.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            data:
              type: object
              properties:
                tableData:
                  type: array
                  items:
                    type: object
                    example: { "columna1": "valor1", "columna2": "valor2" }
                tabla:
                  type: string
                  example: datos
      403:
        description: La tabla solicitada no está permitida.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Tabla no permitida
      400:
        description: Formato no soportado o error en los filtros.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Formato 'xml' no soportado. Use 'json' o 'csv'.
      500:
        description: Error interno en la base de datos o error inesperado.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Error al conectarse a la base de datos <detalle del error>
    """

    args = request.args
    tabla = args.get('tabla')
    limit = args.get('limite')
    offset = int(args.get('offset', 0))
    formato = args.get('formato', 'json')
    orden = args.get('orden', 'desc')
    primarykey = args.get('primarykey', '')

    args_dict = request.args.to_dict()
    not_primary_keys = ['tabla', 'limite', 'offset', 'formato','orden', 'primarykey']

    # Filtrar los argumentos relevantes
    filtered_args = {key: value.split(',') for key, value in args_dict.items() if key not in not_primary_keys}

    # Construir la cláusula WHERE con OR y AND
    where_clauses = []
    params = []

    for key, values in filtered_args.items():
        or_conditions = " OR ".join([f"{key}=%s" for _ in values])
        where_clauses.append(f"({or_conditions})")
        params.extend(values)  # Agregar los valores a los parámetros

    where_clause = ' AND '.join(where_clauses)
    where_clause = f"WHERE {where_clause}" if where_clause else ""

    if tabla not in ALLOWED_TABLES:
        return jsonify({'status': 'fail', 'error': 'Tabla no permitida'}), 403

    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()


        sql_query = f"SELECT * FROM {tabla} {where_clause}"
        if limit is not None:
            sql_query += " LIMIT %s OFFSET %s"
            params.extend([int(limit), offset])

        print("Consulta SQL:", sql_query)
        print("Parámetros:", params)
        cursor.execute(sql_query, params)

        filas = cursor.fetchall()

        respuesta = []
        for fila in filas:
            datos_dict = {key: value for key, value in zip(cursor.column_names, fila)}
            for key, value in datos_dict.items():
                if isinstance(value, decimal.Decimal):
                    datos_dict[key] = float(value)
                elif isinstance(value, (datetime, date)):
                    datos_dict[key] = value.isoformat()
            respuesta.append(datos_dict)

        if formato == 'json':
            json_respuesta = jsonify({
                'status': 'success',
                'data': {
                    'tableData': respuesta,
                    'tabla': tabla
                }
            })
            return json_respuesta, 200, {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'}
        elif formato == 'csv':
            csv_respuesta = generar_csv(respuesta)
            return Response(csv_respuesta, mimetype='text/csv')
        else:
            mensaje_error = f"Formato '{formato}' no soportado. Use 'json' o 'csv'."
            return jsonify({'status': 'fail', 'error': mensaje_error}), 400

    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route("/numeroVariablesProyecto", methods=['GET'])
def numero_variables_por_proyecto():
  try:
    conn = mysql.connector.connect(**config)
    cursor = conn.cursor()
    args = request.args
    id_proyecto = int(args.get("id_proyecto", 1))

    sql_query = f"""
      SELECT disp.id_proyecto, 
        disp.id_dispositivo as id_dispositivo, 
        sens.id_sensor_tipo as id_sensor_tipo, 
        sens.id_sensor as id_sensor,
        ves.idVariable as id_variable
      FROM
        sensores_dev.dispositivos AS disp
      JOIN 
        sensores_dev.sensores_en_dispositivo AS sed ON disp.id_dispositivo = sed.id_dispositivo
      JOIN 
        sensores_dev.sensores AS sens ON sed.id_sensor = sens.id_sensor
      LEFT JOIN 
        sensores_dev.variables_en_sensores AS ves ON sens.id_sensor_tipo = ves.idSensorTipo
      WHERE 
        disp.id_proyecto = %s   
      ORDER BY 
        ves.idVariable ASC
    """

    params = [id_proyecto]
    cursor.execute(sql_query, params)
    filas = cursor.fetchall()
    if len(filas) == 0:
        mensaje_error = f"No hay registros para los filtros solicitados"
        return jsonify({'status': 'fail', 'error': mensaje_error}), 400
    
    num_var= len(filas)
    print(num_var)
    return jsonify(num_var),200
  except Exception as e:
    print(e)
    return None

@app.route('/listarDatosEstructuradosV2', methods=['GET'])
def listar_datos_estructurados_v2():
    """
    Lista los datos estructurados de la tabla "datos" con filtros opcionales.
    ---
    tags:
      - Tablas
    parameters:
      - name: limite
        in: query
        type: integer
        required: false
        description: Número máximo de registros a retornar. Sin límite si no se especifica.
      - name: offset
        in: query
        type: integer
        required: false
        description: Desplazamiento inicial para la consulta. Predeterminado a 0.
      - name: formato
        in: query
        type: string
        required: false
        description: Formato de salida 'json' o 'csv'. Predeterminado a 'json'.
      - name: fecha_inicio
        in: query
        type: string
        format: date
        required: false
        description: Fecha de inicio para filtrar los datos en formato "YYYY-MM-DD".
      - name: fecha_fin
        in: query
        type: string
        format: date
        required: false
        description: Fecha de fin para filtrar los datos en formato "YYYY-MM-DD".
      - name: filtros
        in: query
        type: string
        required: false
        description: Filtros opcionales para columnas específicas en la forma 'columna=valor1,valor2'.
      - name: order_by
        in: query
        type: string
        required: false
        description: Columna por la cual ordenar los resultados. Valores válidos 'fecha', 'id_dato', 'valor', 'codigo_interno', 'timestamp_lectura', 'id_sesion', 'nombre', 'id_proyecto', 'tag'. Predeterminado a 'fecha'.
    responses:
      200:
        description: Datos estructurados obtenidos con éxito.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            data:
              type: object
              properties:
                tableData:
                  type: array
                  items:
                    type: object
                    example: { "fecha": "2024-01-01", "id_sesion": 123, "valor": 45.6, "unidad_medida": "Temperatura (°C)" }
                tabla:
                  type: string
                  example: datos
                totalCount:
                  type: integer
                  example: 100
      400:
        description: No se encontraron registros para los filtros solicitados o error en el formato.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: No hay registros para los filtros solicitados
      403:
        description: La tabla solicitada no está permitida.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Tabla no permitida
      500:
        description: Error interno en la base de datos o error inesperado.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Error al conectarse a la base de datos <detalle del error>
    """

    args = request.args
    tabla = "datos"  # args.get('tabla')  # Nombre de la tabla como parámetro
        
    limit_param = args.get('limite')
    if limit_param is not None:
      limit = int(limit_param)
    else:
      limit = None

    offset = int(args.get('offset', 0))
    formato = args.get('formato', 'json')
    order_by = args.get('order_by', 'fecha')  # Por defecto ordenar por fecha

    print(offset)

    fecha_inicio = args.get('fecha_inicio')
    fecha_fin = args.get('fecha_fin')

    args_dict = request.args.to_dict()
    not_primary_keys = ['tabla', 'limite', 'offset', 'formato', 'fecha_inicio', 'fecha_fin', 'order_by']

    # Filtrar los argumentos relevantes
    filtered_args = {key: value.split(',') for key, value in args_dict.items() if key not in not_primary_keys}

    where_clauses = []
    params = []

    # Mapear order_by a columnas válidas para filtros de fecha
    valid_order_columns_filter = {
        'fecha': 'd.fecha',
        'fecha_insercion': 'd.fecha_insercion',
        'id_dato': 'd.fecha',  # Para id_dato usar fecha por defecto
        'valor': 'd.fecha',    # Para valor usar fecha por defecto
        'codigo_interno': 'd.fecha',  # Para codigo_interno usar fecha por defecto
        'id_sesion': 'd.fecha',       # Para id_sesion usar fecha por defecto
        'id_proyecto': 'd.fecha'      # Para id_proyecto usar fecha por defecto
    }
    
    # Obtener la columna de fecha para filtros según order_by
    fecha_column = valid_order_columns_filter.get(order_by.lower(), 'd.fecha')
    
    # Rango de fechas dinámico según order_by
    if fecha_inicio:
        where_clauses.append(f"({fecha_column} >= %s)")
        params.append(fecha_inicio)
    
    if fecha_fin:
        where_clauses.append(f"({fecha_column} <= %s)")
        params.append(fecha_fin)

    for key, values in filtered_args.items():
        or_conditions = " OR ".join([f"{key}=%s" for _ in values])
        where_clauses.append(f"({or_conditions})")
        params.extend(values)

    where_clause = ' AND '.join(where_clauses)
    if where_clause:
        where_clause = f"WHERE {where_clause}"
    else:
        where_clause = ""

    if tabla not in ALLOWED_TABLES:
        return jsonify({'status': 'fail', 'error': 'Tabla no permitida'}), 403

    
    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        # Adaptar el valor de limit y offset para la tabla estructurada
        id_proyecto = args.get("disp.id_proyecto")
        if id_proyecto is not None:
            result_func = f_numero_variables_por_proyecto(id_proyecto)
            if result_func is not None and isinstance(result_func, dict):
                num_dispositivos = result_func.get("num_dispositivos", 0)
                num_variables_proyecto = result_func.get("num_variables_proyecto", 0)
                num_variables_dispositivo = int(result_func.get("num_variables_dispositivo", 0))
                limit_adaptado = (limit * num_variables_dispositivo) if (limit is not None and limit > 0 and num_variables_dispositivo > 0) else 0
                offset_adaptado = offset * num_variables_dispositivo if offset > 0 and num_variables_dispositivo > 0 else 0
            else:
                print("Error: result_func es None o no es un diccionario")
                num_dispositivos = None
                num_variables_proyecto = None
                num_variables_dispositivo = None
                limit_adaptado = limit if limit is not None else 0
                offset_adaptado = 0
        else:
            num_dispositivos = None
            num_variables_proyecto = None
            num_variables_dispositivo = None
            limit_adaptado = limit if limit is not None else 0
            offset_adaptado = 0
        sql_query = f"""
            SELECT
                d.id_dato,
                d.fecha,
                d.id_sesion,
                d.valor,
                d.fecha_insercion,
                CONCAT(st.modelo, ' [', v.descripcion, ' (', v.unidad, ')]') AS unidad_medida,
                s.descripcion AS sesion_descripcion,
                s.fecha_inicio,
                s.ubicacion,
                disp.id_proyecto,
                disp.codigo_interno,
                disp.descripcion AS dispositivo_descripcion
            FROM
                sensores_dev.datos AS d
            LEFT JOIN
                sensores_dev.variables AS v ON d.id_variable = v.id_variable
            LEFT JOIN
                sensores_dev.sesiones AS s ON d.id_sesion = s.id_sesion
            LEFT JOIN
                sensores_dev.sensores AS sens ON d.id_sensor = sens.id_sensor
            LEFT JOIN
                sensores_dev.sensores_tipo AS st ON sens.id_sensor_tipo = st.id_sensor_tipo
            LEFT JOIN
                sensores_dev.sensores_en_dispositivo AS sed ON sens.id_sensor = sed.id_sensor
            LEFT JOIN
                sensores_dev.dispositivos AS disp ON sed.id_dispositivo = disp.id_dispositivo
            {where_clause}
        """

        # Validar y agregar ORDER BY de forma segura
        valid_order_columns = {
            'fecha': 'd.fecha',
            'fecha_insercion': 'd.fecha_insercion',
            'id_dato': 'd.id_dato',
            'valor': 'd.valor',
            'codigo_interno': 'disp.codigo_interno',
            'id_sesion': 'd.id_sesion',
            'id_proyecto': 'disp.id_proyecto'
        }
        
        safe_order_column = valid_order_columns.get(order_by.lower(), 'd.fecha')
        
        # Manejar valores NULL/vacíos para que aparezcan al final en DESC
        if order_by.lower() == 'fecha_insercion':
            sql_query += f" ORDER BY {safe_order_column} IS NULL, {safe_order_column} DESC"
        else:
            sql_query += f" ORDER BY {safe_order_column} DESC"

        params_sql = params.copy()
        if limit is not None:
          # Solo agrega LIMIT y OFFSET si se entregó el parámetro limite
          sql_query += " LIMIT %s OFFSET %s"
          params_sql.extend([limit_adaptado, offset_adaptado])

        cursor.execute(sql_query, params_sql)
        filas = cursor.fetchall()
        if len(filas) == 0:
            mensaje_error = f"No hay registros para los filtros solicitados"
            return jsonify({'status': 'fail', 'error': mensaje_error}), 400
        
        # Convertir resultados en DataFrame
        respuesta = []
        for fila in filas:
          datos_dict = {key: value for key, value in zip(cursor.column_names, fila)}
          for key, value in datos_dict.items():
            # Convierte id_dato a string explícitamente
            if key == "id_dato" and isinstance(value, int):
              datos_dict[key] = str(value)
            elif isinstance(value, decimal.Decimal):
              datos_dict[key] = float(value)
            elif isinstance(value, (datetime, date)):
              datos_dict[key] = value.isoformat()
          respuesta.append(datos_dict)

        # Crear el DataFrame
        df = pd.DataFrame(respuesta)

        # Rellenar valores nulos (NaN)
        df = df.fillna(value={"id_sesion": "Sin sesión", "fecha_insercion": "", "sesion_descripcion": "", "fecha_inicio": "", "ubicacion": "", "dispositivo_descripcion": ""})

        # Crear la tabla pivotada
        df_pivoted = df.pivot_table(
            index=["fecha", "fecha_insercion", "id_sesion", "sesion_descripcion", "fecha_inicio", "ubicacion", "id_proyecto", "codigo_interno", "dispositivo_descripcion"],
            columns="unidad_medida",
            values="valor",
            aggfunc=list
        ).reset_index()

        # Crear una columna vectorizada con los id_dato concatenados por la columna de ordenamiento
        # Mapear order_by a columnas disponibles en el DataFrame
        valid_groupby_columns = {
            'fecha': 'fecha',
            'fecha_insercion': 'fecha_insercion', 
            'id_sesion': 'id_sesion',
            'codigo_interno': 'codigo_interno',
            'id_proyecto': 'id_proyecto'
        }
        
        groupby_column = valid_groupby_columns.get(order_by.lower(), 'fecha')
        
        id_concat = (
            df.groupby(groupby_column)["id_dato"]
              .apply(lambda s: ', '.join(map(str, s)))
              .reset_index()
              .rename(columns={"id_dato": "id_dato_concatenado"})
        )

        # Unir la columna al DataFrame pivotado
        df_pivoted = df_pivoted.merge(id_concat, on=groupby_column, how="left")

        # Convertir las listas a cadenas separadas por comas
        df_pivoted = df_pivoted.map(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else str(x) if x is not None else "")

        # Calcular total_count antes de aplicar limit y offset (independendiente del limit)
        try:
            codigo_interno=args.get("disp.codigo_interno")
            if not codigo_interno:
                codigos_internos = f_dispositivos_por_proyecto(id_proyecto=id_proyecto)
                total_count = f_numero_mediciones_por_dispositivo(codigo_interno=codigos_internos,fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
            else:
                total_count = f_numero_mediciones_por_dispositivo(codigo_interno=codigo_interno,fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
            #TODO: Que pasa si hay mas sensores en un dispositivo?
            if num_variables_dispositivo and num_variables_dispositivo > 0:
                total_count = int(total_count / num_variables_dispositivo)
            else:
                total_count = len(df_pivoted)
        except Exception as e:
            print(f"Error calculando total_count: {e}")
            total_count = len(df_pivoted)

        # Aplicar limit y offset al DataFrame pivotado
        # if limit > 0:
        #     df_pivoted = df_pivoted.iloc[offset:offset + limit]
        
        # Mapear columnas válidas para el DataFrame pivotado
        valid_pivot_columns = {
            'fecha': 'fecha',
            'fecha_insercion': 'fecha_insercion',
            'id_sesion': 'id_sesion',
            'codigo_interno': 'codigo_interno',
            'id_proyecto': 'id_proyecto'
        }
        
        # Aplicar ORDER BY al DataFrame pivotado
        pivot_order_column = valid_pivot_columns.get(order_by.lower(), 'fecha')
        if pivot_order_column in df_pivoted.columns:
            # Manejar fecha_insercion especialmente para valores vacíos
            if order_by.lower() == 'fecha_insercion':
                # Convertir cadenas vacías a NaT para ordenamiento correcto
                df_pivoted['fecha_insercion_sort'] = pd.to_datetime(df_pivoted['fecha_insercion'], errors='coerce')
                df_pivoted = df_pivoted.sort_values(by='fecha_insercion_sort', ascending=False, na_position='last')
                df_pivoted = df_pivoted.drop('fecha_insercion_sort', axis=1)
            else:
                df_pivoted = df_pivoted.sort_values(by=pivot_order_column, ascending=False)
        else:
            # Fallback a fecha si la columna no existe
            df_pivoted = df_pivoted.sort_values(by="fecha", ascending=False)

        # Formato de respuesta
        if formato == 'json':
            json_response = df_pivoted.to_dict(orient="records")
            json_respuesta = json.dumps({
                'status': 'success',
                'data': {
                    'tableData': json_response,
                    'tabla': tabla,
                    'totalCount': total_count
                }
            }, ensure_ascii=False)
            return json_respuesta, 200, {'Content-Type': 'application/json; charset=utf-8', 'Access-Control-Allow-Origin': '*'}

        elif formato == 'csv':
            return Response(
                stream_with_context(build_csv(df_pivoted)),
                mimetype="text/csv",
                headers={"Content-Disposition": "attachment;filename=output.csv"}
            )

        elif formato == 'xlsx':
            return Response(
                stream_with_context(build_excel(df_pivoted)),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment;filename=output.xlsx"}
            )
        else:
            mensaje_error = f"Formato '{formato}' no soportado. Use 'json' o 'csv'."
            return jsonify({'status': 'fail', 'error': mensaje_error}), 400


    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/listarDatosEstructurados', methods=['GET'])
def listar_datos_estructurados():
    """
    Lista los datos estructurados de la tabla "datos" con filtros opcionales.
    ---
    tags:
      - Tablas
    parameters:
      - name: limite
        in: query
        type: integer
        required: false
        description: Número máximo de registros a retornar. Sin límite si no se especifica.
      - name: offset
        in: query
        type: integer
        required: false
        description: Desplazamiento inicial para la consulta. Predeterminado a 0.
      - name: formato
        in: query
        type: string
        required: false
        description: Formato de salida 'json' o 'csv'. Predeterminado a 'json'.
      - name: fecha_inicio
        in: query
        type: string
        format: date
        required: false
        description: Fecha de inicio para filtrar los datos en formato "YYYY-MM-DD".
      - name: fecha_fin
        in: query
        type: string
        format: date
        required: false
        description: Fecha de fin para filtrar los datos en formato "YYYY-MM-DD".
      - name: filtros
        in: query
        type: string
        required: false
        description: Filtros opcionales para columnas específicas en la forma 'columna=valor1,valor2'.
    responses:
      200:
        description: Datos estructurados obtenidos con éxito.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            data:
              type: object
              properties:
                tableData:
                  type: array
                  items:
                    type: object
                    example: { "fecha": "2024-01-01", "id_sesion": 123, "valor": 45.6, "unidad_medida": "Temperatura (°C)" }
                tabla:
                  type: string
                  example: datos
                totalCount:
                  type: integer
                  example: 100
      400:
        description: No se encontraron registros para los filtros solicitados o error en el formato.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: No hay registros para los filtros solicitados
      403:
        description: La tabla solicitada no está permitida.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Tabla no permitida
      500:
        description: Error interno en la base de datos o error inesperado.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Error al conectarse a la base de datos <detalle del error>
    """

    args = request.args
    tabla = "datos"  # args.get('tabla')  # Nombre de la tabla como parámetro
    limit = int(args.get('limite', 0))
    offset = int(args.get('offset', 0))
    formato = args.get('formato', 'json')

    fecha_inicio = args.get('fecha_inicio')
    fecha_fin = args.get('fecha_fin')

    args_dict = request.args.to_dict()
    not_primary_keys = ['tabla', 'limite', 'offset', 'formato', 'fecha_inicio', 'fecha_fin']

    # Filtrar los argumentos relevantes
    filtered_args = {key: value.split(',') for key, value in args_dict.items() if key not in not_primary_keys}

    where_clauses = []
    params = []

    # Rango de fechas
    if fecha_inicio:
        where_clauses.append("(d.fecha >= %s)")
        params.append(fecha_inicio)
    
    if fecha_fin:
        where_clauses.append("(d.fecha <= %s)")
        params.append(fecha_fin)

    for key, values in filtered_args.items():
        or_conditions = " OR ".join([f"{key}=%s" for _ in values])
        where_clauses.append(f"({or_conditions})")
        params.extend(values)

    where_clause = ' AND '.join(where_clauses)
    where_clause = f"WHERE {where_clause}" if where_clause else ""

    if tabla not in ALLOWED_TABLES:
        return jsonify({'status': 'fail', 'error': 'Tabla no permitida'}), 403

    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        sql_query = f"""
            SELECT
                d.fecha,
                d.id_sesion,
                d.valor,
                CONCAT(st.modelo, ' [', v.descripcion, ' (', v.unidad, ')]') AS unidad_medida,
                s.descripcion AS sesion_descripcion,
                s.fecha_inicio,
                s.ubicacion,
                disp.id_proyecto,
                disp.codigo_interno,
                disp.descripcion AS dispositivo_descripcion
            FROM
                sensores_dev.datos AS d
            LEFT JOIN
                sensores_dev.variables AS v ON d.id_variable = v.id_variable
            LEFT JOIN
                sensores_dev.sesiones AS s ON d.id_sesion = s.id_sesion
            LEFT JOIN
                sensores_dev.sensores AS sens ON d.id_sensor = sens.id_sensor
            LEFT JOIN
                sensores_dev.sensores_tipo AS st ON sens.id_sensor_tipo = st.id_sensor_tipo
            LEFT JOIN
                sensores_dev.sensores_en_dispositivo AS sed ON sens.id_sensor = sed.id_sensor
            LEFT JOIN
                sensores_dev.dispositivos AS disp ON sed.id_dispositivo = disp.id_dispositivo
            {where_clause}
            ORDER BY d.fecha DESC
        """

        cursor.execute(sql_query, params)
        filas = cursor.fetchall()
        if len(filas) == 0:
            mensaje_error = f"No hay registros para los filtros solicitados"
            return jsonify({'status': 'fail', 'error': mensaje_error}), 400
        # Convertir resultados en DataFrame
        respuesta = []
        for fila in filas:
            datos_dict = {key: value for key, value in zip(cursor.column_names, fila)}
            for key, value in datos_dict.items():
                if isinstance(value, decimal.Decimal):
                    datos_dict[key] = float(value)
                elif isinstance(value, (datetime, date)):
                    datos_dict[key] = value.isoformat()
            respuesta.append(datos_dict)

        print("respuesta", respuesta)
        
        df = pd.DataFrame(respuesta)
        df = df.fillna(value={"id_sesion": "Sin sesión", "sesion_descripcion": "", "fecha_inicio": "", "ubicacion": "", "dispositivo_descripcion":""})

        df_pivoted = df.pivot_table(
            index=["fecha", "id_sesion", "sesion_descripcion", "fecha_inicio", "ubicacion", "id_proyecto", "codigo_interno", "dispositivo_descripcion"],
            columns="unidad_medida",
            values="valor",
            aggfunc=list
        ).reset_index()
        

        # Convertir las listas a cadenas separadas por comas
        # df_pivoted = df_pivoted.applymap(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)

        df_pivoted = df_pivoted.applymap(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else str(x) if x is not None else "")


        # Calcular total_count antes de aplicar limit y offset
        total_count = len(df_pivoted)

        # Aplicar limit y offset al DataFrame pivotado
        if limit > 0:
            df_pivoted = df_pivoted.iloc[offset:offset + limit]

        if formato == 'json':
            json_response = df_pivoted.to_dict(orient="records")
            json_respuesta = json.dumps({
                'status': 'success',
                'data': {
                    'tableData': json_response,
                    'tabla': tabla,
                    'totalCount': total_count
                }
            }, ensure_ascii=False)
            return json_respuesta, 200, {'Content-Type': 'application/json; charset=utf-8', 'Access-Control-Allow-Origin': '*'}
        elif formato == 'csv':
            return Response(
                stream_with_context(build_csv(df_pivoted)),
                mimetype="text/csv",
                headers={"Content-Disposition": "attachment;filename=output.csv"}
            )
        elif formato == 'xlsx':
            return Response(
                stream_with_context(build_excel(df_pivoted)),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment;filename=output.xlsx"}
            )

        else:
            mensaje_error = f"Formato '{formato}' no soportado. Use 'json' o 'csv'."
            return jsonify({'status': 'fail', 'error': mensaje_error}), 400

    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/listarUltimasMediciones', methods=['GET'])
def listar_ultimas_mediciones():
    """
    Lista las últimas mediciones estructuradas de la tabla "datos" con filtros opcionales.
    ---
    tags:
      - Tablas
    parameters:
      - name: limite
        in: query
        type: integer
        required: false
        description: Número máximo de registros a retornar. Sin límite si no se especifica.
      - name: offset
        in: query
        type: integer
        required: false
        description: Desplazamiento inicial para la consulta. Predeterminado a 0.
      - name: formato
        in: query
        type: string
        required: false
        description: Formato de salida 'json' o 'csv'. Predeterminado a 'json'.
      - name: fecha_inicio
        in: query
        type: string
        format: date
        required: false
        description: Fecha de inicio para filtrar los datos en formato "YYYY-MM-DD". Se filtra por el campo especificado en order_by (fecha o fecha_insercion).
      - name: fecha_fin
        in: query
        type: string
        format: date
        required: false
        description: Fecha de fin para filtrar los datos en formato "YYYY-MM-DD". Se filtra por el campo especificado en order_by (fecha o fecha_insercion).
      - name: filtros
        in: query
        type: string
        required: false
        description: Filtros opcionales para columnas específicas en la forma 'columna=valor1,valor2'.
      - name: order_by
        in: query
        type: string
        required: false
        description: Campo por el cual ordenar los resultados (fecha, fecha_insercion, id_sesion, codigo_interno, id_proyecto). Predeterminado a 'fecha'. También determina el campo usado para filtros fecha_inicio/fecha_fin.
    responses:
      200:
        description: Últimas mediciones obtenidas con éxito.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            data:
              type: object
              properties:
                tableData:
                  type: array
                  items:
                    type: object
                    example: { "fecha": "2024-01-01", "id_sesion": 123, "valor": 45.6, "unidad_medida": "Temperatura (°C)" }
                tabla:
                  type: string
                  example: datos
      400:
        description: No se encontraron registros para los filtros solicitados o error en el formato.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: No hay registros para los filtros solicitados
      403:
        description: La tabla solicitada no está permitida.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Tabla no permitida
      500:
        description: Error interno en la base de datos o error inesperado.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: Error al conectarse a la base de datos <detalle del error>
    """

    args = request.args
    tabla = "datos"
    limit = int(args.get('limite', 0))
    offset = int(args.get('offset', 0))
    formato = args.get('formato', 'json')
    order_by = args.get('order_by', 'fecha')

    fecha_inicio = args.get('fecha_inicio')
    fecha_fin = args.get('fecha_fin')

    args_dict = request.args.to_dict()
    not_primary_keys = ['tabla', 'limite', 'offset', 'formato', 'fecha_inicio', 'fecha_fin', 'order_by']

    # Filtrar los argumentos relevantes
    filtered_args = {key: value.split(',') for key, value in args_dict.items() if key not in not_primary_keys}

    # Mapear order_by a columnas válidas para filtros de fecha
    valid_order_columns_filter = {
        'fecha': 'd.fecha',
        'fecha_insercion': 'd.fecha_insercion',
        'id_dato': 'd.fecha',  # Para id_dato usar fecha por defecto
        'valor': 'd.fecha',    # Para valor usar fecha por defecto
        'codigo_interno': 'd.fecha',  # Para codigo_interno usar fecha por defecto
        'id_sesion': 'd.fecha',       # Para id_sesion usar fecha por defecto
        'id_proyecto': 'd.fecha'      # Para id_proyecto usar fecha por defecto
    }
    
    # Obtener la columna de fecha para filtros según order_by
    fecha_column = valid_order_columns_filter.get(order_by.lower(), 'd.fecha')

    where_clauses = []
    params = []

    # Rango de fechas dinámico según order_by
    if fecha_inicio:
        where_clauses.append(f"({fecha_column} >= %s)")
        params.append(fecha_inicio)
    
    if fecha_fin:
        where_clauses.append(f"({fecha_column} <= %s)")
        params.append(fecha_fin)

    for key, values in filtered_args.items():
        or_conditions = " OR ".join([f"{key}=%s" for _ in values])
        where_clauses.append(f"({or_conditions})")
        params.extend(values)

    where_clause = ' AND '.join(where_clauses)
    where_clause = f"WHERE {where_clause}" if where_clause else ""

    if tabla not in ALLOWED_TABLES:
        return jsonify({'status': 'fail', 'error': 'Tabla no permitida'}), 403

    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        sql_query = f"""
            SELECT
                d.id_dato,
                d.fecha,
                d.id_sesion,
                d.valor,
                d.fecha_insercion,
                CONCAT(st.modelo, ' [', v.descripcion, ' (', v.unidad, ')]') AS unidad_medida,
                s.descripcion AS sesion_descripcion,
                s.fecha_inicio,
                s.ubicacion,
                disp.id_proyecto,
                disp.codigo_interno,
                disp.descripcion AS dispositivo_descripcion
            FROM
                sensores_dev.datos AS d
            LEFT JOIN
                sensores_dev.variables AS v ON d.id_variable = v.id_variable
            LEFT JOIN
                sensores_dev.sesiones AS s ON d.id_sesion = s.id_sesion
            LEFT JOIN
                sensores_dev.sensores AS sens ON d.id_sensor = sens.id_sensor
            LEFT JOIN
                sensores_dev.sensores_tipo AS st ON sens.id_sensor_tipo = st.id_sensor_tipo
            LEFT JOIN
                sensores_dev.sensores_en_dispositivo AS sed ON sens.id_sensor = sed.id_sensor
            LEFT JOIN
                sensores_dev.dispositivos AS disp ON sed.id_dispositivo = disp.id_dispositivo
            {where_clause}
        """

            # ORDER BY d.fecha DESC

         # Validar y agregar ORDER BY de forma segura
        valid_order_columns = {
            'fecha': 'd.fecha',
            'fecha_insercion': 'd.fecha_insercion',
            'id_dato': 'd.id_dato',
            'valor': 'd.valor',
            'codigo_interno': 'disp.codigo_interno',
            'id_sesion': 'd.id_sesion',
            'id_proyecto': 'disp.id_proyecto'
        }
        
        safe_order_column = valid_order_columns.get(order_by.lower(), 'd.fecha')
        
        # Manejar valores NULL/vacíos para que aparezcan al final en DESC
        if order_by.lower() == 'fecha_insercion':
            sql_query += f" ORDER BY {safe_order_column} IS NULL, {safe_order_column} DESC"
        else:
            sql_query += f" ORDER BY {safe_order_column} DESC"
        
        cursor.execute(sql_query, params)
        filas = cursor.fetchall()
        if len(filas) == 0:
            mensaje_error = f"No hay registros para los filtros solicitados"
            return jsonify({'status': 'fail', 'error': mensaje_error}), 400
        
        # Convertir resultados en DataFrame
        respuesta = []
        for fila in filas:
            datos_dict = {key: value for key, value in zip(cursor.column_names, fila)}
            for key, value in datos_dict.items():
                if isinstance(value, decimal.Decimal):
                    datos_dict[key] = float(value)
                elif isinstance(value, (datetime, date)):
                    datos_dict[key] = value.isoformat()
            respuesta.append(datos_dict)

        print("respuesta", respuesta)
        
        df = pd.DataFrame(respuesta)
        df = df.fillna(value={"id_sesion": "Sin sesión", "sesion_descripcion": "", "fecha_inicio": "", "ubicacion": "", "dispositivo_descripcion":"", "fecha_insercion": ""})

        df_pivoted = df.pivot_table(
            index=["fecha", "fecha_insercion", "id_sesion", "sesion_descripcion", "fecha_inicio", "ubicacion", "id_proyecto", "codigo_interno", "dispositivo_descripcion"],
            columns="unidad_medida",
            values="valor",
            aggfunc=list
        ).reset_index()
        
        # Convertir las listas a cadenas separadas por comas
        df_pivoted = df_pivoted.map(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else str(x) if x is not None else "")

        # Aplicar limit y offset al DataFrame pivotado
        if limit > 0:
            df_pivoted = df_pivoted.iloc[offset:offset + limit]

        if formato == 'json':
            json_response = df_pivoted.to_dict(orient="records")
            json_respuesta = json.dumps({
                'status': 'success',
                'data': {
                    'tableData': json_response,
                    'tabla': tabla
                }
            }, ensure_ascii=False)
            return json_respuesta, 200, {'Content-Type': 'application/json; charset=utf-8', 'Access-Control-Allow-Origin': '*'}
        elif formato == 'csv':
            return Response(
                stream_with_context(build_csv(df_pivoted)),
                mimetype="text/csv",
                headers={"Content-Disposition": "attachment;filename=output.csv"}
            )
        elif formato == 'xlsx':
            return Response(
                stream_with_context(build_excel(df_pivoted)),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment;filename=output.xlsx"}
            )
        else:
            mensaje_error = f"Formato '{formato}' no soportado. Use 'json' o 'csv'."
            return jsonify({'status': 'fail', 'error': mensaje_error}), 400

    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/listarSensores', methods=['GET'])
def listar_sensores():
    args = request.args
    limit = int(args.get('limite', 100))
    offset = int(args.get('offset', 0))
    id_dispositivos_raw = args.get('id_dispositivo')  # Puede ser '1,2,3' o None

    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        sql_query = """
        SELECT 
            sensores.id_sensor,    
            sensores.id_sensor_tipo,
            sensores.numero_serial,
            sensores_tipo.codigo_interno,
            sensores_tipo.marca,    
            sensores_tipo.modelo,
            sensores_tipo.descripcion,
            sensores.variables_usadas
        FROM sensores
        LEFT JOIN sensores_tipo ON sensores.id_sensor_tipo = sensores_tipo.id_sensor_tipo
        LEFT JOIN sensores_en_dispositivo ON sensores.id_sensor = sensores_en_dispositivo.id_sensor
        """

        params = []

        if id_dispositivos_raw:
            # Convertir string '1,2,3' a lista ['1','2','3']
            id_dispositivos = [id_.strip() for id_ in id_dispositivos_raw.split(',') if id_.strip().isdigit()]
            if id_dispositivos:
                placeholders = ','.join(['%s'] * len(id_dispositivos))
                sql_query += f" WHERE sensores_en_dispositivo.id_dispositivo IN ({placeholders})"
                params.extend(id_dispositivos)
        # else no WHERE

        sql_query += " ORDER BY sensores.id_sensor ASC"
        sql_query += " LIMIT %s OFFSET %s"
        params.extend([limit, offset])

        cursor.execute(sql_query, params)
        filas = cursor.fetchall()

        columnas = [
            "Id Sensor",
            "Id Sensor Tipo",
            "N° de Serie",
            "Código Interno",
            "Marca",
            "Modelo",
            "Descripcion",
            "Variables Usadas"
        ]

        respuesta = [dict(zip(columnas, fila)) for fila in filas]

        return jsonify({
            'status': 'success',
            'data': {
                'tableData': respuesta,
                'tabla': 'sensores_combinados'
            }
        }), 200, {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'}

    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/schema', methods=['GET'])
def get_table_schema():
    """
    Obtiene el esquema de una tabla específica en la base de datos.
    ---
    tags:
      - Tablas
    parameters:
      - name: tabla
        in: query
        type: string
        required: true
        description: Nombre de la tabla para la cual se solicita el esquema.
    responses:
      200:
        description: Esquema de la tabla obtenido con éxito.
        schema:
          type: array
          items:
            type: object
            properties:
              Field:
                type: string
                example: "id_sensor"
              Type:
                type: string
                example: "int(11)"
              Null:
                type: string
                example: "NO"
              Key:
                type: string
                example: "PRI"
              Default:
                type: string
                example: null
              Extra:
                type: string
                example: "auto_increment"
              Count:
                type: integer
                example: 100
      400:
        description: Error al obtener el esquema de la tabla.
        schema:
          type: object
          properties:
            error:
              type: string
              example: "Error al obtener esquema: tabla no existe"
      500:
        description: Error interno o de conexión a la base de datos.
        schema:
          type: object
          properties:
            error:
              type: string
              example: "Error de conexión a la base de datos"
    """

    args = request.args
    tabla = args.get('tabla')

    try:
        conn = mysql.connector.connect(**config)
        if not conn:
            return jsonify({"error": "Error de conexión a la base de datos"}), 500
        
        cursor = conn.cursor()
        
        cursor.execute(f"SELECT COUNT(*) FROM {tabla}")
        total_count = cursor.fetchone()[0]

        # Ejecutar una consulta para obtener la información del esquema de la tabla
        cursor.execute(f"DESCRIBE {tabla}")
        schema = cursor.fetchall()
        
        # Transformar el resultado en un formato más legible
        columns = []
        for column in schema:
            column_info = {
                "Field": column[0],
                "Type": column[1],
                "Null": column[2],
                "Key": column[3],
                "Default": column[4],
                "Extra": column[5],
                "Count": total_count
            }
            columns.append(column_info)

        return jsonify(columns), 200

    except mysql.connector.Error as e:
        return jsonify({"error": f"Error al obtener esquema: {e}"}), 400

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/modificarDatos', methods=['PUT'])
def modificar_datos():
    """
    Modifica registros en una tabla específica de la base de datos.
    ---
    tags:
      - Datos
    parameters:
      - name: body
        in: body
        required: true
        description: Datos necesarios para modificar un registro.
        schema:
          type: object
          properties:
            tableName:
              type: string
              example: "sensores"
              description: Nombre de la tabla donde se realizará la modificación.
            primaryKeys:
              type: object
              description: Claves primarias del registro a modificar.
              example: { "id_sensor": 1 }
            formData:
              type: object
              description: Datos que se actualizarán en el registro.
              example: { "descripcion": "Nuevo valor", "estado": "Activo" }
    responses:
      200:
        description: Registro actualizado correctamente.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            message:
              type: string
              example: "1 registro(s) actualizado(s) correctamente"
      403:
        description: La tabla especificada no está permitida.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "Tabla no permitida"
      404:
        description: El registro no fue encontrado o no se realizaron cambios.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "Registro no encontrado o sin cambios"
      500:
        description: Error interno en la base de datos o error inesperado.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "Error al conectarse a la base de datos <detalle del error>"
    """

    data = request.get_json()
    table_name = data.get('tableName')  # Las valores del formulario SIN las primary keys
    primary_keys = data.get('primaryKeys')  
    form_data = data.get('formData')

    concatenated_filter = ' AND'.join([f"{key}={value}" for key, value in primary_keys.items()])
    if concatenated_filter != '':
        concatenated_filter = 'WHERE '+concatenated_filter

    if table_name not in ALLOWED_TABLES:
        return jsonify({'status': 'fail', 'error': 'Tabla no permitida'}), 403
    
    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        # Generar dinámicamente la consulta SQL para actualizar los campos
        set_clause = ", ".join([f"{key} = %s" for key in form_data.keys()])
        valores = list(form_data.values())

        # Construir la consulta de actualización
        sql_query = f"UPDATE {table_name} SET {set_clause} {concatenated_filter}"

        log_query = sql_query % tuple(valores)  # Sustituye los %s por los valores reales
        print("Consulta SQL para depuración:", log_query)

        # Ejecutar la consulta
        cursor.execute(sql_query, valores)
        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({'status': 'fail', 'error': 'Registro no encontrado o sin cambios'}), 404

        return jsonify({'status': 'success', 'message': f'{cursor.rowcount} registro(s) actualizado(s) correctamente'}), 200, {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'}

    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/eliminarDatos', methods=['GET'])
def eliminar_datos():
    """
    Elimina registros de una tabla específica en la base de datos.
    ---
    tags:
      - Datos
    parameters:
      - name: tabla
        in: query
        type: string
        required: true
        description: Nombre de la tabla desde donde se eliminarán los registros.
      - name: filtros
        in: query
        type: string
        required: true
        description: Filtros para identificar los registros a eliminar en la forma 'columna=valor'.
    responses:
      200:
        description: Registro(s) eliminado(s) correctamente.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            message:
              type: string
              example: "1 registro(s) eliminado(s) correctamente"
      403:
        description: Faltan parámetros requeridos o la tabla no está permitida.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "'Se requiere un ID' o 'Tabla no permitida'"
      404:
        description: Registro no encontrado o no se realizaron cambios.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "Registro no encontrado o sin cambios"
      500:
        description: Error interno en la base de datos o error inesperado.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "Error al conectarse a la base de datos <detalle del error>"
    """

    args = request.args
    tabla = args.get('tabla')

    if not tabla:
        return jsonify({'status': 'fail', 'error': 'Se requiere el nombre de la tabla'}), 403

    if tabla not in ALLOWED_TABLES:
        return jsonify({'status': 'fail', 'error': 'Tabla no permitida'}), 403

    id_param = None
    for key in request.args:
        if key != 'tabla':
            id_param = key
            break

    if not id_param:
        return jsonify({'status': 'fail', 'error': 'Se requiere un ID'}), 403

    ids = request.args.get(id_param)
    

    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()
        # Permitir múltiples ids separados por comas
        id_list = [id.strip() for id in ids.split(",") if id.strip().isdigit()]
        if not id_list:
            return jsonify({'status': 'fail', 'error': 'IDs inválidos'}), 403

        placeholders = ','.join(['%s'] * len(id_list))
        sql_query = f"DELETE FROM {tabla} WHERE {id_param} IN ({placeholders})"
        cursor.execute(sql_query, id_list)
        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({'status': 'fail', 'error': 'Registro no encontrado o sin cambios'}), 404

        return jsonify({'status': 'success', 'message': f'{cursor.rowcount} registro(s) eliminado(s) correctamente'}), 200

    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"

        if e.errno == 1451:
            mensaje_error = f"Error: No es posible eliminar el registro pues existe una referencia a este en otra tabla\n{e}"

        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/agregarDatos', methods=['POST'])
def agregar_datos():
    """
    Agrega un nuevo registro a una tabla específica en la base de datos.
    ---
    tags:
      - Datos
    parameters:
      - name: body
        in: body
        required: true
        description: Datos necesarios para insertar un nuevo registro.
        schema:
          type: object
          properties:
            tableName:
              type: string
              example: "sensores"
              description: Nombre de la tabla donde se insertará el registro.
            formData:
              type: object
              description: Datos que se insertarán en el registro.
              example: { "descripcion": "Sensor de temperatura", "estado": "Activo" }
    responses:
      201:
        description: Registro insertado correctamente.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            message:
              type: string
              example: "Registro insertado correctamente"
      403:
        description: La tabla especificada no está permitida.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "Tabla no permitida"
      500:
        description: Error interno en la base de datos o error inesperado.
        schema:
          type: object
          properties:
            status:
              type: string
              example: fail
            error:
              type: string
              example: "Error al conectarse a la base de datos <detalle del error>"
    """

    data = request.get_json()
    table_name = data.get('tableName')  # El nombre de la tabla
    form_data = data.get('formData')  # Los valores del formulario
    if table_name not in ALLOWED_TABLES:
        return jsonify({'status': 'fail', 'error': 'Tabla no permitida'}), 403

    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        # Generar dinámicamente la consulta SQL para insertar los campos
        columns = ", ".join(form_data.keys())
        placeholders = ", ".join(["%s"] * len(form_data))
        valores = list(form_data.values())

        # Construir la consulta de inserción
        sql_query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
        
        log_query = sql_query % tuple(valores)  # Para fines de depuración
        print("Consulta SQL para depuración:", log_query)

        # Ejecutar la consulta
        cursor.execute(sql_query, valores)


        # Intentar obtener el ID de la última fila insertada
        last_inserted_id = cursor.lastrowid if cursor.lastrowid else None

        conn.commit()

        return jsonify({'status': 'success', 'message': 'Registro insertado correctamente', "id": last_inserted_id}), 201, {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'}

    except mysql.connector.Error as e:
        mensaje_error = f"Error al conectarse a la base de datos {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    except Exception as e:
        mensaje_error = f"Error desconocido: {e}"
        print(mensaje_error)
        return jsonify({'status': 'fail', 'error': mensaje_error}), 500

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# Definir el directorio donde se guardarán las imágenes
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Extensiones permitidas para la imagen
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png'}

# Función para verificar las extensiones de archivo permitidas
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
 
# Endpoint para recibir la imagen
@app.route('/agregarImagen', methods=['POST'])
def agregar_imagen():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    if file and allowed_file(file.filename):
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
       
        return jsonify({"message": f"Image saved at {filepath}"}), 200
    else:
        return jsonify({"error": "Invalid file type"}), 400

@app.route('/verImagenes', methods=['GET'])
def ver_imagenes():
    """
    Devuelve una lista de los nombres de las imágenes almacenadas en la carpeta 'uploads'.
    """
    try:
        # Obtener una lista de todos los archivos en la carpeta uploads
        imagenes = os.listdir(app.config['UPLOAD_FOLDER'])
        imagenes = [img for img in imagenes]  # Filtrar solo imágenes
        return jsonify({"imagenes": imagenes}), 200
    except Exception as e:
        return jsonify({"error": f"Error al obtener las imágenes: {e}"}), 500


@app.route('/verImagen/<filename>', methods=['GET'])
def ver_imagen(filename):
    """
    Sirve una imagen desde el servidor para que pueda ser vista en el navegador.
    """
    try:
        # Enviar el archivo solicitado desde la carpeta uploads
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except FileNotFoundError:
        return jsonify({"error": "Imagen no encontrada"}), 404



def generar_csv(data):
    if not data:
        return ''

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    for row in data:
        writer.writerow(row)
    return output.getvalue()

def build_csv(df_pivoted):
    output = io.BytesIO()
    df_pivoted.to_csv(output, index=False, encoding="utf-8-sig")
    output.seek(0)
    for line in output:
        yield line    
    output.close()


# Funcion adaptada para DICTUC
# Crear estos mismos pasos para los demas formatos publicos (sobretodo quitar id dato concatenado)
def build_excel(df_pivoted):
    # Realizar el reemplazo de los encabezados
    df_pivoted.columns = df_pivoted.columns.str.replace('codigo_interno', 'dispositivo', case=False)
    df_pivoted.columns = df_pivoted.columns.str.replace('AM2302 [Grados celcius (°C)]', 'Temperatura Ambiental (°C)', case=False)
    df_pivoted.columns = df_pivoted.columns.str.replace('AM2302 [Humedad (%)]', 'Humedad Ambiental (%)', case=False)
    df_pivoted.columns = df_pivoted.columns.str.replace('CWT-Soil-THC-S [Grados celcius (°C)]', 'Temperatura de Suelo (°C)', case=False)
    df_pivoted.columns = df_pivoted.columns.str.replace('CWT-Soil-THC-S [Humedad relativa del Suelo (% R.H.)]', 'Humedad del Suelo (% R.H.)', case=False)

    # Eliminar la columna con el título 'id_sesion'
    columns_to_drop = [
        'id_sesion', 'sesion_descripcion', 'fecha_inicio', 'ubicacion', 
        'id_proyecto', 'dispositivo_descripcion', 
        'Divisor de Voltaje [Voltaje (V)]', 
        'SIM7600G [Intensidad señal telefónica (Adimensional)]',
        "id_dato_concatenado"
    ]
    df_pivoted = df_pivoted.drop(columns=[col for col in columns_to_drop if col in df_pivoted.columns])

    # Crear el archivo Excel en memoria
    output = io.BytesIO()

    # Crear un nuevo libro de trabajo con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    
    # Convertir el DataFrame a una lista de listas (incluyendo encabezados)
    data = [df_pivoted.columns.to_list()] + df_pivoted.values.tolist()

    # Escribir los datos en la hoja de trabajo
    for row in data:
        ws.append(row)
    
    # Crear un objeto de tabla en openpyxl
    table = Table(displayName="DatosTabla", ref=ws.dimensions)

    # Establecer el estilo de la tabla (opcional)
    style = TableStyleInfo(
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Añadir la tabla a la hoja
    ws.add_table(table)

    # Establecer estilo azul para la tabla y alternar las filas con fondo negro
    blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    dark_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Aplicar los estilos de fondo azul y alternar filas negras
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
        for cell in row:
            if row_idx % 2 == 0:  # Fila par (blanca)
                cell.fill = light_fill
            else:  # Fila impar (negra)
                cell.fill = blue_fill

            # Colorear la cabecera (si es necesario)
            # if row_idx == -1:
            #     cell.fill = light_fill

    # Guardar el archivo en memoria
    wb.save(output)

    # Mover el puntero al principio para leer el archivo
    output.seek(0)
    
    # Generar el archivo para que se pueda enviar como respuesta
    for line in output:
        yield line
        
    output.close()



# FUNCIONES
def f_numero_mediciones_por_dispositivo(codigo_interno, fecha_inicio=None, fecha_fin=None):
  """Devuelve el número total de mediciones aplicando los filtros.

  - `codigo_interno` puede ser None, string, comma-separated string o lista de strings.
  - La función siempre devuelve un entero con el total (suma sobre códigos si se pasan varios).
  """
  conn = None
  cursor = None
  try:
    conn = mysql.connector.connect(**config)
    cursor = conn.cursor()

    where_clauses = []
    params = []

    # Normalizar codigo_interno a lista si se entrega
    codigo_list = None
    if codigo_interno:
      if isinstance(codigo_interno, list):
        codigo_list = [str(c).strip() for c in codigo_interno if c is not None and str(c).strip() != '']
      elif isinstance(codigo_interno, str) and ',' in codigo_interno:
        codigo_list = [c.strip() for c in codigo_interno.split(',') if c.strip() != '']
      else:
        codigo_list = [str(codigo_interno).strip()]

    if codigo_list:
      placeholders = ','.join(['%s'] * len(codigo_list))
      where_clauses.append(f"(disp.codigo_interno IN ({placeholders}))")
      params.extend(codigo_list)

    if fecha_inicio:
      where_clauses.append("(d.fecha >= %s)")
      params.append(fecha_inicio)

    if fecha_fin:
      where_clauses.append("(d.fecha <= %s)")
      params.append(fecha_fin)

    where_clause = ' AND '.join(where_clauses)
    where_clause = f"WHERE {where_clause}" if where_clause else ""

    # Hacemos una sola consulta que cuente todas las filas que cumplan los filtros
    count_query = f"""
      SELECT COUNT(*)
      FROM
        sensores_dev.datos AS d
      LEFT JOIN
        sensores_dev.sensores AS sens ON d.id_sensor = sens.id_sensor
      LEFT JOIN
        sensores_dev.sensores_en_dispositivo AS sed ON sens.id_sensor = sed.id_sensor
      LEFT JOIN
        sensores_dev.dispositivos AS disp ON sed.id_dispositivo = disp.id_dispositivo
      {where_clause}
    """
    cursor.execute(count_query, params)
    row = cursor.fetchone()
    total_count = int(row[0]) if row and row[0] is not None else 0
    return total_count

  except mysql.connector.Error as e:
    print(f"Error en base de datos al contar mediciones: {e}")
    return 0
  except Exception as e:
    print(f"Error desconocido al contar mediciones: {e}")
    return 0
  finally:
    try:
      if cursor is not None:
        cursor.close()
    except Exception:
      pass
    try:
      if conn is not None and conn.is_connected():
        conn.close()
    except Exception:
      pass


def f_dispositivos_por_proyecto(id_proyecto):
  try:
    conn = mysql.connector.connect(**config)
    cursor = conn.cursor()
    sql_query = "SELECT DISTINCT codigo_interno FROM sensores_dev.dispositivos WHERE id_proyecto = %s"
    params = [id_proyecto]
    cursor.execute(sql_query, params)
    rows = cursor.fetchall()


    return [row[0] for row in rows]
  except mysql.connector.Error as e:
    print(f"Error DB en f_numero_mediciones_por_proyecto: {e}")
    return []
  except Exception as e:
    print(f"Error desconocido en f_numero_mediciones_por_proyecto: {e}")
    return []
  finally:
    if 'conn' in locals() and conn.is_connected():
      cursor.close()
      conn.close()

def f_numero_variables_por_proyecto(id_proyecto):
    try:
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()

        sql_query = """
            SELECT disp.id_proyecto, 
                disp.id_dispositivo as id_dispositivo, 
                sens.id_sensor_tipo as id_sensor_tipo, 
                sens.id_sensor as id_sensor,
                ves.idVariable as id_variable
            FROM
                sensores_dev.dispositivos AS disp
            JOIN 
                sensores_dev.sensores_en_dispositivo AS sed ON disp.id_dispositivo = sed.id_dispositivo
            JOIN 
                sensores_dev.sensores AS sens ON sed.id_sensor = sens.id_sensor
            LEFT JOIN 
                sensores_dev.variables_en_sensores AS ves ON sens.id_sensor_tipo = ves.idSensorTipo
            WHERE 
                disp.id_proyecto = %s   
            ORDER BY 
                ves.idVariable ASC
        """

        params = [id_proyecto]
        cursor.execute(sql_query, params)
        filas = cursor.fetchall()

        if len(filas) == 0:
            mensaje_error = f"No hay registros para los filtros solicitados"
            print(mensaje_error)
            return {
                "num_dispositivos": 0,
                'num_variables_proyecto': 0,
                "num_variables_dispositivo": 0, 
            }

        # Crear el JSON con los datos
        resultado = []
        for fila in filas:
            variable = {
                'id_proyecto': fila[0],
                'id_dispositivo': fila[1],
                'id_sensor_tipo': fila[2],
                'id_sensor': fila[3],
                'id_variable': fila[4]
            }
            resultado.append(variable)

        dispositivos_unicos = {item["id_dispositivo"] for item in resultado}

        # Contar el número de dispositivos únicos
        num_dispositivos = len(dispositivos_unicos)
        print(num_dispositivos)
        # Crear el JSON con el número total de variables
        num_var = len(filas)
        result = {
            "num_dispositivos": num_dispositivos,
            'num_variables_proyecto': num_var,
            "num_variables_dispositivo": len(filas) / num_dispositivos, 
        }
        print(result['num_dispositivos'])
        return result

    except Exception as e:
        print(f"Error en f_numero_variables_por_proyecto: {e}")
        return {
            "num_dispositivos": 0,
            'num_variables_proyecto': 0,
            "num_variables_dispositivo": 0, 
        }

    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


# Alertas




if __name__ == "__main__":
    app.run(host='0.0.0.0', port=8084)
    # from eventlet import wsgi
    # import eventlet
    # eventlet.monkey_patch()
    # socketio.run(app, host='0.0.0.0', port=8084, debug=True)

